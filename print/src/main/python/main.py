import itertools
import re
import sys  # sys нужен для передачи argv в QApplication
from pathlib import Path
from typing import Optional

import openpyxl
from fbs_runtime.application_context.PyQt5 import ApplicationContext
from openpyxl.styles import Alignment
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5 import QtWidgets
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem

from print.src.main.python.qt import mainwindow  # Это наш конвертированный файл дизайна


class ExampleApp(QtWidgets.QMainWindow, mainwindow.Ui_MainWindow):
    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна
        self.buttonBox.hide()
        self.btnBrowse.clicked.connect(
            self.browse_folder
        )  # Выполнить функцию browse_folder
        # при нажатии кнопки

        self.buttonBox.accepted.connect(self.process_file)
        # self.buttonBox.rejected.connect(self.process_file)

        self.file: Optional[Path] = None

    def browse_folder(self):
        self.buttonBox.hide()
        self.tableWidget.clear()  # На случай, если в списке уже есть элементы
        dialog = QtWidgets.QFileDialog(self)
        (filename, _) = dialog.getOpenFileName(
            dialog,
            "Выберите таблицу",
            options=QFileDialog.DontUseNativeDialog,
            filter="Таблицы (*.xls *.xlsx)",
        )

        self._handle_browse_result(filename)

    def _handle_browse_result(self, filename: Optional[str]):
        if not filename:
            return

        self.file = Path(filename).resolve()

        self.wb = openpyxl.open(self.file)
        self.ws = self.wb.active

        self.buttonBox.setEnabled(True)
        self.setWindowTitle(f'Active - {self.file.name}')
        self._load_table_widget(self.ws)
        self.buttonBox.show()

    def _load_table_widget(self, ws: Worksheet, row_start: int = 2):
        headers = [
            cell.value for cell in itertools.takewhile(lambda x: bool(x), ws['1'])
        ]
        self.tableWidget.setColumnCount(len(headers))
        self.tableWidget.setRowCount(ws.max_row - row_start + 1)

        self.tableWidget.setHorizontalHeaderLabels(headers)
        self.tableWidget.setVerticalHeaderLabels(
            map(str, range(row_start, ws.max_row + 1))
        )

        data = ws.iter_rows(min_row=row_start)
        for x, row in enumerate(data):
            if row[0].value is None:
                break

            for y, cell in enumerate(row):
                val = cell.value
                if val is None:
                    continue
                item = QTableWidgetItem(str(val))
                self.tableWidget.setItem(x, y, item)

        self.range_start.setText('A2')
        self.range_end.setText(f'A{ws.max_row}')

    def process_file(self):
        if not self.file:
            return

        row_start, col_start = coordinate_to_tuple(self.range_start.text())
        row_end, col_end = coordinate_to_tuple(self.range_end.text())

        for row in range(row_start, row_end + 1):

            cell_content = self.ws.cell(row, col_start).value
            if not cell_content:
                continue

            if sheet_name_match := re.search(
                r'^(\d*/[-\d\w]*?(?:[a-zа-я](?=[А-ЯA-Z])|$))', cell_content
            ):
                head = sheet_name_match.group(1)
                sheet_name = f'Печать {head.replace("/", " ")}'
                print(f'Adding {sheet_name}...')
                new_ws = self.wb.create_sheet(sheet_name)

                target_cell_content = f"{head}\n" + re.sub(
                    r'((ш(ирина)?|в(ысота)?)\s+)?(\d+(\s*[*xх]\s*)?)+',
                    "\n\g<0>",
                    cell_content[len(head) :],
                )

                target_row, target_col = coordinate_to_tuple(self.target_cell.text())
                new_cell = new_ws[self.target_cell.text()]
                new_cell.value = target_cell_content

                new_cell.alignment = Alignment(wrap_text=True, vertical='top')
                new_ws.column_dimensions[get_column_letter(target_col)].width = 20
                new_ws.row_dimensions[target_row].height = 300

                print('Sheet processed')

        self._save_workbook()

    def _save_workbook(self):
        edited_file = (self.file.parent / f'Обработка_{self.file.stem}').with_suffix(
            self.file.suffix
        )
        print(f'Saving to {edited_file}...')
        self.wb.save(edited_file)

        QDesktopServices.openUrl(QUrl.fromLocalFile(str(edited_file)))


def main():
    appctxt = ApplicationContext()  # 1. Instantiate ApplicationContext

    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно

    exit_code = appctxt.app.exec_()  # 2. Invoke appctxt.app.exec_()
    sys.exit(exit_code)


if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()
