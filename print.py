import itertools
import re
import sys  # sys нужен для передачи argv в QApplication
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.utils.units import cm_to_dxa, points_to_pixels, cm_to_EMU, EMU_to_inch
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5 import QtWidgets
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QColor, QDesktopServices
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem

from qt import mainwindow  # Это наш конвертированный файл дизайна

WIDTH_CM_LIST = [2.11, 1.21, 1.08, 0.94, 1.19, 0.94, 6.06, 10.35]
HEIGHT_CM_LIST = [0.51] * 8 + [0.35, 4.71]
MARGIN_X_CM = 0.64
MARGIN_Y_CM = 1.91



class ExampleApp(QtWidgets.QMainWindow, mainwindow.Ui_MainWindow):
    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна
        self.buttonBox.hide()
        self.btnBrowse.clicked.connect(self.browse_folder)  # Выполнить функцию browse_folder
        # при нажатии кнопки

        self.buttonBox.accepted.connect(self.process_file)
        # self.buttonBox.rejected.connect(self.process_file)

        self.file: Optional[Path] = None

    def browse_folder(self):
        self.buttonBox.hide()
        self.tableWidget.clear()  # На случай, если в списке уже есть элементы
        dialog = QtWidgets.QFileDialog(self)
        (filename, _) = dialog.getOpenFileName(
            dialog,
            "Выберите таблицу",
            options=QFileDialog.DontUseNativeDialog,
            filter="Таблицы (*.xls *.xlsx)",
        )

        self._handle_browse_result(filename)

    def _handle_browse_result(self, filename: Optional[str]):
        if not filename:
            return

        self.file = Path(filename)

        self.wb = openpyxl.open(self.file)
        self.ws = self.wb.active

        self.buttonBox.setEnabled(True)
        self.setWindowTitle(f'Active - {self.file.name}')
        self._load_table_widget(self.ws)
        self.buttonBox.show()

    def _load_table_widget(self, ws: Worksheet, row_start: int = 2):
        headers = [cell.value for cell in itertools.takewhile(lambda x: bool(x), ws['1'])]
        self.tableWidget.setColumnCount(len(headers))
        self.tableWidget.setRowCount(ws.max_row - row_start + 1)

        self.tableWidget.setHorizontalHeaderLabels(headers)
        self.tableWidget.setVerticalHeaderLabels(map(str, range(row_start, ws.max_row + 1)))

        data = ws.iter_rows(min_row=row_start)
        for x, row in enumerate(data):
            if row[0].value is None:
                break

            for y, cell in enumerate(row):
                val = cell.value
                if val is None:
                    continue
                item = QTableWidgetItem(str(val))
                self.tableWidget.setItem(x, y, item)

        self.range_start.setText('A2')
        self.range_end.setText(f'A{ws.max_row}')

    def process_file(self):
        if not self.file:
            return

        warn_color = 'ffef5c'
        warning_fill = PatternFill(start_color=warn_color, end_color=warn_color, fill_type='solid')
        q_color = QColor.fromRgb(int(f'0x{warn_color}', base=16))

        row_start, col_start = coordinate_to_tuple(self.range_start.text())
        row_end, col_end = coordinate_to_tuple(self.range_end.text())

        for row in range(row_start, row_end + 1):

            cell_content = self.ws.cell(row, col_start).value
            if not cell_content:
                continue

            # fmt: off
            if sheet_name_match := re.search(
                r'^(\d*/[-\d\w]*?'
                    r'(?:[a-zа-я]'
                        r'(?=[А-ЯA-Z])|(?=\s*артикул)|$'
                    r')'
                r')', cell_content, flags=re.X
            ):
                # fmt: on
                head = sheet_name_match.group(1)
                sheet_name = f'Печать {head.replace("/", " ")}'
                print(f'Adding {sheet_name}...')
                new_ws = self.wb.create_sheet(sheet_name)
                self._format_new_worksheet(new_ws)

                target_cell_content = f"{head}\n" + re.sub(
                    r'((артикул|ш(ирина)?|в(ысота)?)\s+)?(\d+(\s*[*xх]\s*)?)+',
                    "\n\g<0>",
                    cell_content[len(head) :],
                )

                new_cell = new_ws[self.target_cell.text()]
                new_cell.value = target_cell_content

                print('Sheet processed')
            else:
                print(f'No match for content {cell_content}')
                self.ws.cell(row, col_start).fill = warning_fill
                self.tableWidget.item(*to_widget_coords(row, col_start)).setBackground(q_color)

        self._save_workbook()

    def _format_new_worksheet(self, new_ws):
        new_cell = new_ws[self.target_cell.text()]

        new_cell.alignment = Alignment(wrap_text=True, vertical='top')
        new_cell.font = Font('Calibri', 12, b=True)

        for col, col_width_cm in enumerate(WIDTH_CM_LIST, start=1):
            new_ws.column_dimensions[get_column_letter(col)].width = (
                points_to_pixels(cm_to_dxa(col_width_cm) / 20, dpi=72) / 7
            )

        for row, row_height_cm in enumerate(HEIGHT_CM_LIST, start=1):
            new_ws.row_dimensions[row].height = points_to_pixels(
                cm_to_dxa(row_height_cm) / 20, dpi=72
            )

        new_ws.print_area = f'A1:{get_column_letter(len(WIDTH_CM_LIST))}{len(HEIGHT_CM_LIST)}'
        new_ws.page_margins.top = new_ws.page_margins.bottom = EMU_to_inch(cm_to_EMU(MARGIN_Y_CM))
        new_ws.page_margins.left = new_ws.page_margins.right = EMU_to_inch(cm_to_EMU(MARGIN_X_CM))

    def _save_workbook(self):
        edited_file = (self.file.parent / f'Обработка_{self.file.stem}').with_suffix(
            self.file.suffix
        )
        print(f'Saving to {edited_file}...')
        self.wb.save(edited_file)

        QDesktopServices.openUrl(QUrl.fromLocalFile(str(edited_file)))


def to_widget_coords(row, col):
    HEADER = 1
    return row - 1 - HEADER, col - 1


def main():
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение


if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()
